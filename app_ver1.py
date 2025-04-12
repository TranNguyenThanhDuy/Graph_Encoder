import serial
import time
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
from datetime import datetime
import threading
import queue
import os
from openpyxl import load_workbook, Workbook

# Cấu hình
COM_PORT = 'COM3'  # Thay đổi nếu cần
BAUD_RATE = 9600
EXCEL_FILE = 'encoder_data_ver1.xlsx'

# Hàng đợi cho dữ liệu UART
data_queue = queue.Queue()

# Bộ đệm trong bộ nhớ
timestamps = []
values = []

# Khóa để đồng bộ ghi Excel
excel_lock = threading.Lock()

# Biến toàn cục cho serial
serial_port = None

# Thời gian cập nhật Excel định kỳ
last_excel_read = 0
EXCEL_READ_INTERVAL = 1  # Đọc Excel mỗi 5 giây

# Khởi tạo kết nối UART
def connect_serial():
    global serial_port
    while True:
        try:
            serial_port = serial.Serial(COM_PORT, BAUD_RATE, timeout=1)
            print(f"Đã kết nối tới {COM_PORT}")
            return serial_port
        except Exception as e:
            print(f"Lỗi kết nối: {e}. Thử kết nối lại sau 2 giây...")
            time.sleep(2)

# Đọc dữ liệu từ UART
def read_serial():
    global serial_port
    while True:
        try:
            if serial_port is None or not serial_port.is_open:
                serial_port = connect_serial()
            if serial_port.in_waiting > 0:
                data = serial_port.readline().decode('utf-8').strip()
                timestamp = datetime.now().strftime('%H:%M:%S')
                value = float(data)
                data_queue.put((timestamp, value))
        except Exception as e:
            print(f"Lỗi đọc dữ liệu: {e}")
            if serial_port is not None and serial_port.is_open:
                serial_port.close()
            serial_port = connect_serial()

# Tạo hoặc khởi tạo file Excel
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Timestamp', 'Encoder_Value'])
        wb.save(EXCEL_FILE)
        print(f"Đã tạo file Excel mới: {EXCEL_FILE}")
    # Đọc dữ liệu ban đầu vào bộ đệm
    try:
        df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
        if not df.empty:
            timestamps.extend(df['Timestamp'].tolist())
            values.extend(df['Encoder_Value'].tolist())
            print(f"Đã đọc {len(df)} dòng từ Excel vào bộ đệm")
    except Exception as e:
        print(f"Lỗi khởi tạo Excel: {e}")

# Ghi dữ liệu vào Excel
def write_to_excel(timestamp, value):
    with excel_lock:
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.append([timestamp, value])
            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"Lỗi ghi Excel: {e}")
            wb = Workbook()
            ws = wb.active
            ws.append(['Timestamp', 'Encoder_Value'])
            ws.append([timestamp, value])
            wb.save(EXCEL_FILE)

# Luồng ghi dữ liệu vào Excel và cập nhật bộ đệm
def write_excel_thread():
    while True:
        if not data_queue.empty():
            timestamp, value = data_queue.get()
            timestamps.append(timestamp)
            values.append(value)
            write_to_excel(timestamp, value)
            if len(timestamps) > 50:
                timestamps.pop(0)
                values.pop(0)

# Hàm khởi tạo đồ thị
def init_graph():
    line.set_data([], [])
    return line,

# Hàm cập nhật đồ thị từ bộ đệm
def update_graph(frame):
    global last_excel_read
    current_time = time.time()

    # Đọc từ Excel định kỳ (mỗi 5 giây)
    if current_time - last_excel_read > EXCEL_READ_INTERVAL:
        try:
            with excel_lock:
                df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
            if not df.empty:
                timestamps[:] = df['Timestamp'].tolist()[-50:]  # Cập nhật toàn bộ bộ đệm
                values[:] = df['Encoder_Value'].tolist()[-50:]
                print(f"Đã cập nhật bộ đệm từ Excel: {timestamps[-1]}, {values[-1]}")
            last_excel_read = current_time
        except Exception as e:
            print(f"Lỗi đọc Excel trong đồ thị: {e}")

    # Cập nhật đồ thị từ bộ đệm
    if timestamps and values:
        line.set_data(range(len(values)), values)
        ax.set_xticks(range(len(timestamps)))
        ax.set_xticklabels(timestamps, rotation=45)
        if values:
            ax.set_ylim(min(values) - 0.1 * abs(min(values)), max(values) + 0.1 * abs(max(values)))
        ax.relim()
        ax.autoscale_view(scaley=False)
    return line,

# Khởi tạo
init_excel()
serial_port = connect_serial()

# Chạy luồng đọc và ghi
serial_thread = threading.Thread(target=read_serial, daemon=True)
serial_thread.start()
excel_thread = threading.Thread(target=write_excel_thread, daemon=True)
excel_thread.start()

# Cấu hình đồ thị
fig, ax = plt.subplots(figsize=(10, 6))
line, = ax.plot([], [], '-b', label='Encoder Value')
ax.set_title('Real-time Encoder Data from Excel')
ax.set_xlabel('Time')
ax.set_ylabel('Encoder Value')
ax.legend()

# Tạo animation
ani = FuncAnimation(fig, update_graph, init_func=init_graph, blit=True, interval=1000)
plt.tight_layout()
plt.show()

# Đóng cổng serial
if serial_port is not None and serial_port.is_open:
    serial_port.close()